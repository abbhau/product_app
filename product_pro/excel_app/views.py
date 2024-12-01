import pandas as pd
from django.http import HttpResponse
from django.db.models import Prefetch
from openpyxl.utils import get_column_letter
from .models import School, Student, Teacher, Class, Subject
from django.db import connection
import time

def export_school_data_to_excel(request):
     # Start time for performance measurement
    start_time = time.time()

    # Clear any previous queries to get a fresh count
    connection.queries.clear()
    # Prefetch related data to minimize queries (only selecting necessary fields)
    students = Student.objects.only('id', 'name', 'age', 'grade_level')
    teachers = Teacher.objects.only('id', 'name', 'subject', 'hire_date')
    classes = Class.objects.only('id', 'name', 'schedule_time')
    subjects = Subject.objects.only('id', 'name', 'description')

    # Query all schools (prefetching only necessary related data)
    schools = School.objects.prefetch_related(
        Prefetch('students', queryset=students),
        Prefetch('teachers', queryset=teachers),
        Prefetch('classes', queryset=classes),
        Prefetch('subjects', queryset=subjects)
    ).iterator(chunk_size=1000)  # Use chunk_size here

    # Create an HTTP response with the Excel file as an attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=school_data.xlsx'

    # Use openpyxl to write the data to the Excel file
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        # Create the workbook and the worksheet
        workbook = writer.book
        sheet = workbook.create_sheet("School Data")

        # Column headers for the sheet (as per your requirement)
        headers = [
            'ID', 'NAME', 'ADDRESS', 'CITY', 'STATE',
            'STUDENT NAME', 'STUDENT AGE', 'STUDENT GRADE',
            'TEACHER NAME', 'TEACHER SUBJECT', 'TEACHER HIRE DATE',
            'CLASS NAME', 'CLASS SCHEDULE', 'SUBJECT NAME', 'SUBJECT DESCRIPTION'
        ]
        sheet.append(headers)

        # Initialize previous school ID to track when to insert an empty row
        previous_school_id = None
        row_idx = 2  # Start from row 2 to leave space for headers

        # Iterate through schools and write data row by row
        for school in schools:
            students_count = len(school.students.all())
            teachers_count = len(school.teachers.all())
            classes_count = len(school.classes.all())
            subjects_count = len(school.subjects.all())
            max_length = max(
                students_count, teachers_count, classes_count, subjects_count
            )

            # Check if the school ID has changed, and add an empty row if it has
            if school.id != previous_school_id and row_idx != 2:
                # Append an empty row to separate schools
                sheet.append([''] * len(headers))  # Adds an empty row
                row_idx += 1  # Move to the next row

            # Iterate over the related data (students, teachers, classes, subjects)
            for i in range(max_length):
                # Create the row data
                row = [
                    school.id,
                    school.name,
                    school.address,
                    school.city,
                    school.state,
                    school.students.all()[i].name if i < students_count else None,
                    school.students.all()[i].age if i < students_count else None,
                    school.students.all()[i].grade_level if i < students_count else None,
                    school.teachers.all()[i].name if i < teachers_count else None,
                    school.teachers.all()[i].subject if i < teachers_count else None,
                    school.teachers.all()[i].hire_date if i < teachers_count else None,
                    school.classes.all()[i].name if i < classes_count else None,
                    school.classes.all()[i].schedule_time if i < classes_count else None,
                    school.subjects.all()[i].name if i < subjects_count else None,
                    school.subjects.all()[i].description if i < subjects_count else None
                ]

                # Write the row to the sheet
                sheet.append(row)
                row_idx += 1  # Move to the next row

            # Update the previous school ID to track changes
            previous_school_id = school.id

        # Adjust column widths based on content size (max of header length or content length)
        for col_idx, column in enumerate(headers, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(column)  # Start with the length of the header

            # Find the maximum length of the content in the column
            for row in sheet.iter_rows(min_row=2, max_row=row_idx):  # Skip header row
                cell_value = str(row[col_idx - 1].value if row[col_idx - 1].value is not None else '')
                max_length = max(max_length, len(cell_value))

            # Set the width of the column to be the greater of the header length or content length
            sheet.column_dimensions[column_letter].width = max_length + 2  # Add padding for aesthetics
    # Measure execution time
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"Execution time: {execution_time:.4f} seconds")

    # Print out SQL queries
    print(f"Number of queries: {len(connection.queries)}")
    for query in connection.queries:
        print(query['sql'])
    return response
